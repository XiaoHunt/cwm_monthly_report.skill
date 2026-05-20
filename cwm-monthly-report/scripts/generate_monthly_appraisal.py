#!/usr/bin/env python
"""Extract Raven monthly work items and fill CyweeMotion appraisal workbooks."""

from __future__ import annotations

import argparse
import calendar
import json
import math
import re
import sys
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - environment guidance
    raise SystemExit(
        "Missing dependency: openpyxl. Run this script with the Codex bundled Python "
        "runtime or install openpyxl in the active Python environment."
    ) from exc


CONTEXT_COLUMNS = {
    "num",
    "project_name",
    "function",
    "algo_ver",
}

HEADER_ALIASES = {
    "num": {"num", "no", "number"},
    "project_name": {"projectname", "project"},
    "function": {"function", "functions"},
    "algo_ver": {"algover", "algorithmversion", "algoversion"},
    "quantity": {"quantity", "qty"},
    "issues": {"issues", "issue"},
    "status": {"status"},
    "owners": {"owners", "owner"},
    "completion_date": {"completiondate", "completedate", "finishdate"},
    "notes": {"notes", "note"},
}

REQUIRED_COLUMNS = {"project_name", "issues", "status", "owners", "completion_date"}
FILL_CELLS = {
    "main_description": "F5",
    "main_score": "G5",
    "execution_description": "F9",
    "execution_score": "G9",
    "collaboration_description": "F10",
    "collaboration_score": "G10",
}


@dataclass
class WorkItem:
    row: int
    num: str
    project_name: str
    function: str
    algo_ver: str
    quantity: str
    issues: str
    status: str
    owners: str
    completion_date: str
    notes: str
    completion_dates: list[str]
    issue_dates: list[str]
    matched_by: list[str]
    is_closed: bool


def normalize_header(value: Any) -> str:
    text = stringify(value).lower()
    return re.sub(r"[^a-z0-9]+", "", text)


def stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def canonical_header(value: Any) -> str | None:
    normalized = normalize_header(value)
    for canonical, aliases in HEADER_ALIASES.items():
        if normalized in aliases:
            return canonical
    return None


def parse_period(value: str | None) -> tuple[date, date]:
    if not value:
        raise ValueError("Could not infer appraisal period. Provide --period YYYY-MM.")
    match = re.search(r"(20\d{2})\s*年\s*(1[0-2]|0?[1-9])\s*月", value)
    if not match:
        match = re.search(r"(20\d{2})[-/.](1[0-2]|0?[1-9])", value)
    if not match:
        raise ValueError(f"Could not infer appraisal period from: {value!r}")
    year = int(match.group(1))
    month = int(match.group(2))
    return month_bounds(year, month)


def parse_period_arg(period: str | None, template_path: Path) -> tuple[date, date]:
    if period:
        match = re.fullmatch(r"(20\d{2})-(0[1-9]|1[0-2])", period)
        if not match:
            raise ValueError("--period must use YYYY-MM format")
        return month_bounds(int(match.group(1)), int(match.group(2)))

    wb = load_workbook(template_path, data_only=False)
    ws = wb.active
    fragments = [stringify(ws[cell].value) for cell in ("D2", "E2", "F2")]
    return parse_period(" ".join(fragment for fragment in fragments if fragment))


def month_bounds(year: int, month: int) -> tuple[date, date]:
    last_day = calendar.monthrange(year, month)[1]
    return date(year, month, 1), date(year, month, last_day)


def parse_cell_dates(value: Any) -> list[date]:
    if value is None or value == "":
        return []
    if isinstance(value, datetime):
        return [value.date()]
    if isinstance(value, date):
        return [value]
    return parse_text_dates(str(value))


def parse_text_dates(text: str) -> list[date]:
    results: list[date] = []

    for match in re.finditer(r"(?<!\d)(20\d{2})\s*[./-]\s*(\d{1,2})\s*[./-]\s*(\d{1,2})", text):
        append_valid_date(results, match.group(1), match.group(2), match.group(3))

    for match in re.finditer(r"(?<!\d)(20\d{2})(\d{1,2})(\d{2})(?!\d)", text):
        append_valid_date(results, match.group(1), match.group(2), match.group(3))

    for match in re.finditer(r"(?<!\d)(20\d{2})(\d{1,2})(\d{1})(?!\d)", text):
        year, month, day = match.groups()
        append_valid_date(results, year, month, day)

    return sorted(set(results))


def append_valid_date(results: list[date], year_text: str, month_text: str, day_text: str) -> None:
    try:
        parsed = date(int(year_text), int(month_text), int(day_text))
    except ValueError:
        return
    results.append(parsed)


def in_period(values: list[date], start: date, end: date) -> bool:
    return any(start <= value <= end for value in values)


def find_header_map(ws: Any) -> dict[str, int]:
    best_row = None
    best_map: dict[str, int] = {}
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10)):
        current: dict[str, int] = {}
        for cell in row:
            canonical = canonical_header(cell.value)
            if canonical:
                current[canonical] = cell.column
        if len(current) > len(best_map):
            best_row = row[0].row
            best_map = current
    missing = REQUIRED_COLUMNS - set(best_map)
    if missing:
        raise ValueError(f"Raven worksheet is missing required columns: {sorted(missing)}")
    best_map["_header_row"] = best_row or 1
    return best_map


def extract_items(workbook_path: Path, start: date, end: date) -> list[WorkItem]:
    wb = load_workbook(workbook_path, data_only=False)
    if "Raven" not in wb.sheetnames:
        raise ValueError("Project workbook does not contain a 'Raven' worksheet")
    ws = wb["Raven"]
    header_map = find_header_map(ws)
    header_row = header_map.pop("_header_row")
    last_context = {key: "" for key in CONTEXT_COLUMNS}
    items: list[WorkItem] = []

    for row_number in range(header_row + 1, ws.max_row + 1):
        row_values: dict[str, Any] = {}
        for key, col in header_map.items():
            row_values[key] = ws.cell(row_number, col).value

        if not any(stringify(value) for value in row_values.values()):
            continue

        for key in CONTEXT_COLUMNS:
            value = stringify(row_values.get(key))
            if value:
                last_context[key] = value
            else:
                row_values[key] = last_context.get(key, "")

        issues = stringify(row_values.get("issues"))
        if not issues:
            continue

        completion_dates = parse_cell_dates(row_values.get("completion_date"))
        issue_dates = parse_text_dates(issues)
        matched_by: list[str] = []
        if in_period(completion_dates, start, end):
            matched_by.append("completion_date")
        if in_period(issue_dates, start, end):
            matched_by.append("issue_date")
        if not matched_by:
            continue

        status = stringify(row_values.get("status"))
        item = WorkItem(
            row=row_number,
            num=stringify(row_values.get("num")),
            project_name=stringify(row_values.get("project_name")),
            function=stringify(row_values.get("function")),
            algo_ver=stringify(row_values.get("algo_ver")),
            quantity=stringify(row_values.get("quantity")),
            issues=issues,
            status=status,
            owners=stringify(row_values.get("owners")),
            completion_date=stringify(row_values.get("completion_date")),
            notes=stringify(row_values.get("notes")),
            completion_dates=[value.isoformat() for value in completion_dates],
            issue_dates=[value.isoformat() for value in issue_dates],
            matched_by=matched_by,
            is_closed=status.upper() == "CLOSE",
        )
        items.append(item)
    return items


def clamp(value: int, low: int, high: int) -> int:
    return max(low, min(high, value))


def compute_scores(items: list[WorkItem]) -> dict[str, Any]:
    issue_count = len(items)
    project_count = len({item.project_name for item in items if item.project_name})
    closed_count = sum(1 for item in items if item.is_closed)
    open_count = issue_count - closed_count
    weighted_points = issue_count + 1.5 * project_count + 0.5 * closed_count + 0.75 * open_count

    if weighted_points < 15:
        main_score = 85
    elif weighted_points < 25:
        main_score = 86
    elif weighted_points < 35:
        main_score = 87
    elif weighted_points < 45:
        main_score = 88
    else:
        main_score = 89

    return {
        "issue_count": issue_count,
        "project_count": project_count,
        "closed_count": closed_count,
        "open_count": open_count,
        "weighted_points": round(weighted_points, 2),
        "recommended_scores": {
            "main_score": main_score,
            "execution_score": clamp(3 + math.floor(weighted_points / 18), 3, 6),
            "collaboration_score": clamp(3 + math.floor(project_count / 6), 3, 4),
        },
    }


def summarize_projects(items: list[WorkItem]) -> list[dict[str, Any]]:
    grouped: dict[str, list[WorkItem]] = defaultdict(list)
    for item in items:
        grouped[item.project_name or "(未命名项目)"].append(item)

    summary = []
    for project, project_items in sorted(grouped.items(), key=lambda entry: (-len(entry[1]), entry[0])):
        owners = sorted({item.owners for item in project_items if item.owners})
        statuses = Counter(item.status or "(blank)" for item in project_items)
        summary.append(
            {
                "project_name": project,
                "issue_count": len(project_items),
                "closed_count": sum(1 for item in project_items if item.is_closed),
                "open_count": sum(1 for item in project_items if not item.is_closed),
                "owners": owners,
                "statuses": dict(statuses),
                "sample_issues": [shorten(item.issues, 180) for item in project_items[:3]],
            }
        )
    return summary


def shorten(text: str, limit: int) -> str:
    collapsed = re.sub(r"\s+", " ", text).strip()
    return collapsed if len(collapsed) <= limit else collapsed[: limit - 1] + "..."


def build_summary(workbook_path: Path, template_path: Path, start: date, end: date, items: list[WorkItem]) -> dict[str, Any]:
    scores = compute_scores(items)
    open_items = [item for item in items if not item.is_closed]
    return {
        "source_workbook": str(workbook_path),
        "template_workbook": str(template_path),
        "period": {"start": start.isoformat(), "end": end.isoformat()},
        "counts": {key: scores[key] for key in ("issue_count", "project_count", "closed_count", "open_count")},
        "weighted_points": scores["weighted_points"],
        "recommended_scores": scores["recommended_scores"],
        "project_summary": summarize_projects(items),
        "open_or_in_progress_items": [asdict(item) for item in open_items],
        "items": [asdict(item) for item in items],
    }


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def write_summary_md(path: Path, summary: dict[str, Any]) -> None:
    lines = [
        "# CWM Monthly Report Extraction Summary",
        "",
        f"- Source workbook: `{summary['source_workbook']}`",
        f"- Template workbook: `{summary['template_workbook']}`",
        f"- Period: {summary['period']['start']} to {summary['period']['end']}",
        f"- Issues: {summary['counts']['issue_count']}",
        f"- Projects: {summary['counts']['project_count']}",
        f"- Closed: {summary['counts']['closed_count']}",
        f"- Open/in progress: {summary['counts']['open_count']}",
        f"- Weighted points: {summary['weighted_points']}",
        f"- Recommended scores: main {summary['recommended_scores']['main_score']}, "
        f"execution {summary['recommended_scores']['execution_score']}, "
        f"collaboration {summary['recommended_scores']['collaboration_score']}",
        "",
        "## Project Summary",
        "",
    ]
    for project in summary["project_summary"]:
        lines.append(
            f"- {project['project_name']}: {project['issue_count']} items, "
            f"{project['closed_count']} closed, {project['open_count']} open, "
            f"owners: {', '.join(project['owners']) or 'N/A'}"
        )
        for sample in project["sample_issues"]:
            lines.append(f"  - {sample}")

    if summary["open_or_in_progress_items"]:
        lines.extend(["", "## Open Or In-Progress Items", ""])
        for item in summary["open_or_in_progress_items"]:
            lines.append(
                f"- Row {item['row']} | {item['project_name']} | owner={item['owners'] or 'N/A'} | "
                f"status={item['status'] or '(blank)'} | {shorten(item['issues'], 220)}"
            )

    lines.extend(["", "## All Matched Items", ""])
    for item in summary["items"]:
        lines.append(
            f"- Row {item['row']} | {item['project_name']} | status={item['status'] or '(blank)'} | "
            f"owner={item['owners'] or 'N/A'} | matched={','.join(item['matched_by'])} | "
            f"{shorten(item['issues'], 220)}"
        )

    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def command_extract(args: argparse.Namespace) -> int:
    workbook_path = Path(args.workbook)
    template_path = Path(args.template)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    start, end = parse_period_arg(args.period, template_path)
    items = extract_items(workbook_path, start, end)
    summary = build_summary(workbook_path, template_path, start, end, items)

    write_json(output_dir / "summary.json", summary)
    write_summary_md(output_dir / "summary.md", summary)
    print(f"Wrote {output_dir / 'summary.json'}")
    print(f"Wrote {output_dir / 'summary.md'}")
    print(
        "Counts: "
        f"{summary['counts']['issue_count']} items, "
        f"{summary['counts']['project_count']} projects, "
        f"{summary['counts']['closed_count']} closed, "
        f"{summary['counts']['open_count']} open"
    )
    return 0


def validate_draft(draft: dict[str, Any]) -> None:
    missing = [key for key in FILL_CELLS if key not in draft]
    if missing:
        raise ValueError(f"Draft JSON is missing required fields: {missing}")
    for key in ("main_score", "execution_score", "collaboration_score"):
        if not isinstance(draft[key], int):
            raise ValueError(f"{key} must be an integer")
    for key in ("main_description", "execution_description", "collaboration_description"):
        value = draft[key]
        if not isinstance(value, str) or not value.strip():
            raise ValueError(f"{key} must be a non-empty string")
        if looks_like_encoding_loss(value):
            raise ValueError(
                f"{key} appears to contain encoding-loss placeholders ('?'). "
                "Regenerate draft.json as UTF-8 before filling the workbook."
            )


def looks_like_encoding_loss(value: str) -> bool:
    question_count = value.count("?")
    has_cjk = any("\u4e00" <= char <= "\u9fff" for char in value)
    return "????" in value or (question_count >= 10 and not has_cjk)


def fill_template(template_path: Path, draft_path: Path, output_dir: Path) -> Path:
    draft = json.loads(draft_path.read_text(encoding="utf-8"))
    validate_draft(draft)

    wb = load_workbook(template_path, data_only=False)
    ws = wb.active
    for key, coord in FILL_CELLS.items():
        ws[coord].value = draft[key]

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{template_path.stem}_filled{template_path.suffix}"
    wb.save(output_path)
    return output_path


def command_fill(args: argparse.Namespace) -> int:
    output_path = fill_template(Path(args.template), Path(args.draft_json), Path(args.output_dir))
    print(f"Wrote {output_path}")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    subparsers = parser.add_subparsers(dest="command", required=True)

    extract = subparsers.add_parser("extract", help="Extract Raven work items and write summary files")
    extract.add_argument("--workbook", required=True, help="Project tracking workbook containing a Raven worksheet")
    extract.add_argument("--template", required=True, help="CyweeMotion appraisal workbook")
    extract.add_argument("--output-dir", required=True, help="Directory for summary.json and summary.md")
    extract.add_argument("--period", help="Override appraisal month as YYYY-MM")
    extract.set_defaults(func=command_extract)

    fill = subparsers.add_parser("fill", help="Fill a copy of the appraisal workbook from draft JSON")
    fill.add_argument("--template", required=True, help="CyweeMotion appraisal workbook")
    fill.add_argument("--draft-json", required=True, help="Draft JSON with descriptions and scores")
    fill.add_argument("--output-dir", required=True, help="Directory for the filled workbook")
    fill.set_defaults(func=command_fill)
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        return args.func(args)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
